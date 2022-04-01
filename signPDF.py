# -*- coding: utf-8 -*-
""" Python signPDF Wizard
"""
from PyPDF4 import PdfFileWriter, PdfFileReader
from PyPDF4.utils import PdfReadError
import io
import os
import sys
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import date
import locale
import ghostscript
import win32print
import win32api
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog as fd
import yaml
from openpyxl import load_workbook
import atexit
from tkinter import StringVar
import textwrap




def read_yaml(file_path):
    try:
        with open(file_path, "r") as f:
            return yaml.safe_load(f)
    except:
        print("Error opening config file.")
    finally:
        f.close()


config = read_yaml('./conf.yaml')
if not config:
    config = [
    {
     'CONFIG_FILE' : 'conf.yaml',
     'default_signed_filename' : 'voucher_signed.pdf',
     'remote_host' : 'https://westerngovernorsuniversity.sharepoint.com/:f:/r/sites/AmazonCareerChoice/',
     'remote_dir' : 'Shared Documents/2022 Vouchers/',
     'remote_file' : 'Amazon Career Choice Tracker_V1.0.xlsx',
     'local_file' : 'sign_pdf_data.xlsx',
     }
    ]

    
def write_yaml():
    with open(config[0]['CONFIG_FILE'], "w") as f:
        yaml.dump(config, f)
        f.close()

outputvar = ""
        
# update our config variables upon program exit
def exit_handler():
    write_yaml()
    
atexit.register(exit_handler)



# Return printer name to use. If -printer is set it will return this value only if value match with available
# printers list. Return a error if -printer not in list. If no printer specified, retrieve default printer and return
# its name. Sometime default printer is on USELESS_PRINTER list so first printer return by getAvailablePrinters() is
# return. If no printer is return display an error.
def getPrinter():
    default_printer = win32print.GetDefaultPrinter()
    return default_printer

# Use GhostScript API to silent print .pdf and .ps. Use win32api to print .txt. Return a error if printing failed or
# file ext doesn't match.
def printFile(filepath):
    try:
        if os.path.splitext(filepath)[1] in [".pdf", ".ps"]:
            args = [
                "-dPrinted", "-dBATCH", "-dNOSAFER", "-dNOPAUSE", "-dNOPROMPT"
                                                                  "-q",
                "-dNumCopies#1",
                "-sDEVICE#mswinpr2",
                f'-sOutputFile#"%printer%{getPrinter()}"',
                f'"{filepath}"'
            ]

            encoding = locale.getpreferredencoding()
            args = [a.encode(encoding) for a in args]
            ghostscript.Ghostscript(*args)
        elif os.path.splitext(filepath)[1] in [".txt"]:
            # '"%s"' % enable to encapsulate string with quote
            win32api.ShellExecute(0, "printto", '"%s"' % filepath, '"%s"' % getPrinter(), ".", 0)
        return True

    except:
        print("Printing error for file: ", '"%s"' % filepath, "| Printer: ", '"%s"' % getPrinter())
        return False

    

def open_file():
    filetypes = (('PDF Files', '*.pdf'),('All Files', '*.*'))
    filepath = fd.askopenfilename(filetypes=filetypes)
    #text.insert('1.0', filepath.readlines())
    print("Decrypting "+filepath+"...")
    #text.pack()
    
    
    if not os.path.isfile(filepath) and not os.path.exists(filepath):
        print("Path provided is not a file path.")
        sys.exit(2)
    if not printFile(filepath):
        sys.exit(1)




def display_sheets(file):
    wb = load_workbook(file)
    return wb.sheetnames

    
    
courses = {}

def load_excel_data():
    global courses
    rfile = config[0]['local_file']
    wb = load_workbook(filename = rfile)
    sheet=wb['Courses']
    
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
        courseid = row[0]
        course_data = {
            'coursename' : row[1],
            'credits' : row[2]
            }
        courses[courseid] = course_data

    wb.close()
    
load_excel_data()   
 


def fill_excel_values():
    
    ## Load in excel data
    rfile = config[0]['local_file']
    wb = load_workbook(filename = rfile)
    sheet = wb['Sign']
#    coursename = sheet["E2"].value
    courseid = sheet["A2"].value
#    credits = sheet["F2"].value
    tuition = sheet["B2"].value
    start_date = sheet["C2"].value
    end_date = sheet["D2"].value
    courseid_entry.delete(0, 'end')
    courseid_entry.insert(0, courseid)
    coursename_entry.delete(0, 'end')
    coursename_entry.insert(0, get_coursename(courseid))
    credits_entry.delete(0, 'end')
    credits_entry.insert(0, get_credits(courseid))
    tuition_entry.delete(0, 'end')
    tuition_entry.insert(0, "${:,.2f}".format(tuition))
    start_date_entry.delete(0, 'end')
    start_date_entry.insert(0, start_date.strftime('%m/%d/%Y'))
    end_date_entry.delete(0, 'end')
    end_date_entry.insert(0, end_date.strftime('%m/%d/%Y'))
    print("Values loaded from excel.")
    wb.close()


def get_coursename(courseid):
    global courses
    if courseid in courses:
        return courses[courseid]['coursename']
    
def get_credits(courseid):
    global courses
    if courseid in courses:
        return courses[courseid]['credits']




def sign_file(choice):
    global courses_list
    ## file chooser for exising file to sign
    filetypes = (('PDF Files', '*.pdf'),('All Files', '*.*'))
    infilepath = fd.askopenfilename(filetypes=filetypes)
    print("Signing "+infilepath+"...")
    
    memStream = io.BytesIO()
    can = canvas.Canvas(memStream, pagesize=letter)
    
    """ Draw initials on page 2 """
    if choice == 'name' or choice == 'both':
        can.drawString(180, 364, initial_entry.get())
    
    """ Draw course data on page 2 """
    if choice == 'course' or choice == 'both':
        #### Perform drawing on page 2
        pdfmetrics.registerFont(TTFont('Arial', 'C:\\WINDOWS\\FONTS\\ARIAL.TTF'))
        wrapper = textwrap.TextWrapper(width=28)
        
        ### Determine if we have one or more courses to draw ###
        if len(courses_list)>0:
            ### Courses to draw from listbox
            # calculate total tuition
            total_tuition = 0
            for course in courses_list:
                tuition = course[1].replace('$','') # strip out dollar signs
                total_tuition += float(tuition)
            num = 0
            ### distance between course lines
            spacer = 29.5
            for course in courses_list:
                coursename_list = wrapper.wrap(text=course[4])
                start_y = 292-num*spacer
                if len(coursename_list) > 2:
                    start_y=292-num*spacer
                    start_x=70
                    can.setFont('Arial', 6)
                elif len(coursename_list) == 2:
                    start_x=70
                    start_y=287-num*spacer
                    can.setFont('Arial', 8)
                else:
                    start_x=70
                    can.setFont('Arial', 10)
                    start_y=280-num*spacer
                for i in range(len(coursename_list)):
                    can.drawString(start_x, start_y-(i*10), coursename_list[i])
                ### courseID, tuition, start, end, course name, credits ###
                can.setFont('Arial', 12)
                can.drawString(215, 280-num*spacer, course[0])
                can.drawString(310, 280-num*spacer, course[5])
                can.setFont('Arial', 10)
                can.drawString(365, 280-num*spacer, course[2])
                can.drawString(425, 280-num*spacer, course[3])
                can.setFont('Arial', 12)
                can.drawString(490, 280-num*spacer, "${:,.2f}".format(float(course[1].replace('$',''))))
                num = num + 1
            can.drawString(445, 48, "${:,.2f}".format(total_tuition))
        else:
            ### Listbox is empty, draw from entry boxes
            coursename_list = wrapper.wrap(text=coursename_entry.get())
            start_y = 292
            if len(coursename_list) > 2:
                start_y=292
                start_x=70
                can.setFont('Arial', 6)
            elif len(coursename_list) == 2:
                start_x=70
                start_y=287
                can.setFont('Arial', 8)
            else:
                start_x=70
                can.setFont('Arial', 10)
                start_y = 280
            for i in range(len(coursename_list)):
                can.drawString(start_x, start_y-(i*10), coursename_list[i])
            can.setFont('Arial', 12)
            can.drawString(215, 280, courseid_entry.get())
            can.drawString(310, 280, credits_entry.get())
            can.setFont('Arial', 10)
            can.drawString(365, 280, start_date_entry.get())
            can.drawString(425, 280, end_date_entry.get())
            can.setFont('Arial', 12)
            can.drawString(490, 280, "${:,.2f}".format(float(tuition_entry.get().replace('$',''))))
            can.drawString(445, 48, "${:,.2f}".format(float(tuition_entry.get().replace('$',''))))
            
    
    # finish drawing page 2 and move to page 3
    can.showPage()
    
    """ Draw name fields on page 3 """
    if choice == 'name' or choice == 'both':
        can.drawString(180, 561, user_entry.get())
        can.drawString(180, 528, email_entry.get())
        today = date.today()
        can.drawString(420, 494, today.strftime("%m/%d/%Y"))
        pdfmetrics.registerFont(TTFont('Edwardian', 'C:\\WINDOWS\\FONTS\\ITCEDSCR.TTF'))
        can.setFont('Edwardian', 20)
        can.drawString(200, 494, user_entry.get())
    # finish page 3
    can.showPage()
    # save and close output pdf document
    can.save()
    
    #move to the beginning of the StringIO buffer
    memStream.seek(0)
    
    # create a new PDF with Reportlab
    new_pdf = PdfFileReader(memStream)
    
    #open file for binary reading
    try:
        infile = open(infilepath, 'rb')
        existing_pdf = PdfFileReader(infile,strict=False)
    except FileNotFoundError as e:
        del memStream
        del can
        print("Error opening file:",e)
        return False
    
    #list of pages made from merging existing document with
    #signature fields created in memory
    pages = []
    try:
        for i in range(existing_pdf.numPages):
            # create a new page object instead of simply getting a reference
            page = existing_pdf.getPage(i)
            if i==1:
                page.mergePage(new_pdf.getPage(0))
            elif i==2:
                page.mergePage(new_pdf.getPage(1))
            pages.append(page)
    except PdfReadError as e:
        print("Error opening file:",e)
    finally:
        #infile.close()
        del new_pdf
        del existing_pdf
    
    # ask for output file name
    outfilepath = infilepath
    while outfilepath == infilepath:
        outfilepath = fd.asksaveasfilename(filetypes=filetypes)
        if ".pdf" not in outfilepath:
            outfilepath = outfilepath+".pdf"
        
    #open new pdf file in memory for writer for output
    outfile = PdfFileWriter()
    outStream = open(outfilepath, 'wb')

    try:
        #place all of the pages from memory into output pdf
        for i in pages:
            outfile.addPage(i)
        outfile.write(outStream)
        print("File signed, saved to",outfilepath+", operation completed.")
    except ValueError as e:
        print("Unable to output pages:",e)
    finally:
        infile.close()
        outStream.close()
        del outfile
    
    
def sign_file_with_course():
    sign_file('course')

def sign_file_with_name():
    sign_file('name')

def sign_file_with_both():
    sign_file('both')
    
def write_frame(fr, text):
    fr.insert(tk.END, "\n"+text)
    
def update_courseid(sv):
    sv.set(sv.get().upper())
    if sv.get() in courses:
        coursename_entry.delete(0, 'end')
        coursename_entry.insert(0, get_coursename(sv.get()))
        credits_entry.delete(0, 'end')
        credits_entry.insert(0, get_credits(sv.get()))
        
courses_list = []
        
def add_course():
    ### push course into course_list, then update listbox to match
    ### courseID, tuition, start, end, course name, credits ###
    course = [courseid_entry.get(), tuition_entry.get(), start_date_entry.get(), end_date_entry.get(), coursename_entry.get(), credits_entry.get()]
    global courses_list
    if len(courses_list)<4:
        courses_list.append(course)
    rebuild_listbox()
    
def delete_course():
    selected = courses_listbox.curselection()
    global courses_list
    for x in selected:
        del courses_list[x]
    rebuild_listbox()
    
def clear_courses():
    global courses_list
    courses_list = []
    rebuild_listbox()
    
def rebuild_listbox():
    global courses_list
    courses_listbox.delete(0, courses_listbox.size())
    for course in courses_list:
        courses_listbox.insert(tk.END, course[0]+' - '+"${:,.2f}".format(float(course[1].replace('$','')))+' - '+course[5]+'CR')
    

""" Tkinter dialog window functions """

root = tk.Tk()
root.title("WGU Voucher Wizard")
root.resizable(False, False)
root.geometry('500x280')
frame_left = tk.Frame(root)
frame_left.pack(side='top')

initial_entry = tk.Entry(frame_left, width=30)
initial_entry.delete(0,'end')
initial_entry.insert(0, config[0]['initials'])
initial_entry.grid(column=0,row=1)

user_entry = tk.Entry(frame_left, width=30)
user_entry.delete(0, 'end')
user_entry.insert(0, config[0]['name'])
user_entry.grid(column=0,row=2)

email_entry = tk.Entry(frame_left, width=30)
email_entry.delete(0, 'end')
email_entry.insert(0, config[0]['email'])
email_entry.grid(column=0,row=3)


decrypt_button = ttk.Button(
    frame_left,
    text='Decrypt File',
    command=open_file
    )
decrypt_button.grid(column=0,row=0)
#decrypt_button.grid(column=0, row=1, sticky='ws', padx=10, pady=10)
sign_name_button = ttk.Button(
    frame_left,
    text='Sign file with name data',
    command=sign_file_with_name
    )
sign_name_button.grid(column=0,row=4)
#sign_button.grid(column=0, row=2, sticky='ws', padx=10, pady=10)

sign_excel_button = ttk.Button(
    frame_left,
    text='Sign file with course data',
    command=sign_file_with_course
    )
sign_excel_button.grid(column=2,row=7)

sign_both_button = ttk.Button(
    frame_left,
    text='Sign file with all data',
    command=sign_file_with_both
    )
sign_both_button.grid(column=1,row=8)

fill_button = ttk.Button(
    frame_left,
    text='Get values from excel',
    command=fill_excel_values
    )
fill_button.grid(column=2,row=0)


### CourseID Label ###
label1 = tk.Label(frame_left, width=10, text="CourseID")
label1.grid(column=1,row=1)
sv = StringVar()
sv.trace("w", lambda name, index, mode, sv=sv: update_courseid(sv))
### CourseID Entry Box ###
courseid_entry = tk.Entry(frame_left, width=30, textvariable=sv)
courseid_entry.grid(column=2,row=1)
### Tuition Label ###
label2 = tk.Label(frame_left, width=10, text="Tuition")
label2.grid(column=1,row=2)
### Tuition Entry Box ###
tuition_entry = tk.Entry(frame_left, width=30)
tuition_entry.grid(column=2,row=2)
### Start Date Label ###
label3 = tk.Label(frame_left, width=10, text="Start Date")
label3.grid(column=1,row=3)
### Start Date Entry Box ###
start_date_entry = tk.Entry(frame_left, width=30)
start_date_entry.grid(column=2,row=3)
### End Date Label ###
label4 = tk.Label(frame_left, width=10, text="End Date")
label4.grid(column=1,row=4)
### End Date Entry Box ###
end_date_entry = tk.Entry(frame_left, width=30)
end_date_entry.grid(column=2,row=4)
### Course Name Label ###
label5 = tk.Label(frame_left, width=10, text="Course Name")
label5.grid(column=1,row=5)
### Course Name Entry Box ###
coursename_entry = tk.Entry(frame_left, width=30)
coursename_entry.grid(column=2,row=5)
### Credits Label ###
label6 = tk.Label(frame_left, width=10, text="Credits")
label6.grid(column=1,row=6)
### Credits Entry Box ###
credits_entry = tk.Entry(frame_left, width=30)
credits_entry.grid(column=2,row=6)


frame_courses = ttk.Frame(frame_left)
frame_courses.grid(column=2,row=8)


### Add course button ###
add_course_button = ttk.Button(
    frame_courses,
    text='+',
    command=add_course,
    width='5'
    )
add_course_button.grid(column=0,row=0)

### Delete course button ###
delete_course_button = ttk.Button(
    frame_courses,
    text='-',
    command=delete_course,
    width='5'
    )
delete_course_button.grid(column=1,row=0)

### Clear courses button ###
clear_courses_button = ttk.Button(
    frame_courses,
    text='Clear',
    command=clear_courses,
    width='7'
    )
clear_courses_button.grid(column=2,row=0)

### Courses Label ###
label7 = tk.Label(frame_left, width=10, text="Courses")
label7.grid(column=1,row=9)
courses_listbox = tk.Listbox(
    frame_left,
    height='4',
    selectmode='SINGLE',
    width='30'
    )
courses_listbox.grid(column=2,row=9)

fill_excel_values()

root.mainloop()


